from flask_restful import Resource
from flask import send_file
from models import Expense
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO

class ReportResource(Resource):
    def get(self, type, param):
        if type == 'day':
            # Daily tallies JSON
            try:
                target_date = datetime.strptime(param, '%Y-%m-%d').date()
                day_data = Expense.get_day_tallies(target_date)
                return day_data, 200
            except ValueError:
                return {'error': 'Invalid date format. Use YYYY-MM-DD'}, 400
        elif type == 'daily':
            # Daily report - Excel
            try:
                target_date = datetime.strptime(param, '%Y-%m-%d').date()
                day_data = Expense.get_day_tallies(target_date)

                # Create Excel workbook
                wb = Workbook()
                ws = wb.active
                ws.title = f"Daily Report {param}"

                # Headers
                headers = ['Date', 'Type', 'Card No', 'Procedure/Expense', 'Payment Method', 'Invoice Source', 'Amount']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                row = 2
                # Collections
                for collection in day_data['collections']:
                    ws.cell(row=row, column=1, value=param)
                    ws.cell(row=row, column=2, value='Collection')
                    ws.cell(row=row, column=3, value=collection['card_no'])
                    ws.cell(row=row, column=4, value=collection['procedure'])
                    ws.cell(row=row, column=5, value=collection['payment_method'])
                    ws.cell(row=row, column=6, value=collection.get('invoice_source', ''))
                    ws.cell(row=row, column=7, value=collection['amount'])
                    row += 1

                # Expenses
                for expense in day_data['expenses']:
                    ws.cell(row=row, column=1, value=param)
                    ws.cell(row=row, column=2, value='Expense')
                    ws.cell(row=row, column=3, value='')
                    ws.cell(row=row, column=4, value=expense['expense_name'])
                    ws.cell(row=row, column=5, value=expense['payment_method'])
                    ws.cell(row=row, column=6, value='')
                    ws.cell(row=row, column=7, value=expense['amount'])
                    row += 1

                # Totals
                row += 1
                ws.cell(row=row, column=6, value='Totals:')
                row += 1
                totals = day_data['totals']
                ws.cell(row=row, column=5, value='Cash Total')
                ws.cell(row=row, column=7, value=totals['cash_total'])
                row += 1
                ws.cell(row=row, column=5, value='Mpesa Total')
                ws.cell(row=row, column=7, value=totals['mpesa_total'])
                row += 1
                ws.cell(row=row, column=5, value='Till Total')
                ws.cell(row=row, column=7, value=totals['till_total'])
                row += 1
                ws.cell(row=row, column=5, value='Invoice Total')
                ws.cell(row=row, column=7, value=totals['invoice_total'])
                row += 1
                ws.cell(row=row, column=5, value='Card Total')
                ws.cell(row=row, column=7, value=totals['card_total'])
                row += 1
                ws.cell(row=row, column=5, value='Mobile Money Total')
                ws.cell(row=row, column=7, value=totals['mobile_money_total'])
                row += 1
                ws.cell(row=row, column=5, value='Gross Collections')
                ws.cell(row=row, column=7, value=totals['gross_collections'])
                row += 1
                ws.cell(row=row, column=5, value='Total Expenses')
                ws.cell(row=row, column=7, value=totals['total_expenses'])
                row += 1
                ws.cell(row=row, column=5, value='Net Total')
                ws.cell(row=row, column=7, value=totals['net_total'])

                # Save to BytesIO
                bio = BytesIO()
                wb.save(bio)
                bio.seek(0)

                return send_file(bio, as_attachment=True, download_name=f"daily_report_{param}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            except ValueError:
                return {'error': 'Invalid date format. Use YYYY-MM-DD'}, 400
        elif type == 'monthly':
            # Monthly report - JSON
            try:
                year, month_num = map(int, param.split('-'))
                month_data = Expense.get_month_tallies(month_num, year)
                return month_data, 200
            except ValueError:
                return {'error': 'Invalid month format. Use YYYY-MM'}, 400
        else:
            return {'error': 'Invalid type. Use day, daily, or monthly'}, 400