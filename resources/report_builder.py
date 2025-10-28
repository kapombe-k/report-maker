# report_builder.py
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
from .styles import ExcelStyles


class DailyReportBuilder:
    """Builder class for generating formatted daily Excel reports"""

    def __init__(self, date_str, day_data):
        self.date_str = date_str
        self.day_data = day_data
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Daily Report"
        self.current_row = 1

    def build(self):
        """Build the complete report"""
        self._add_title()
        self._add_collections_section()
        self._add_expenses_section()
        self._add_totals_section()
        self._set_column_widths()
        return self._save_to_bytes()

    def _add_title(self):
        """Add report title"""
        self.ws.merge_cells(f"A{self.current_row}:D{self.current_row}")
        title_cell = self.ws.cell(row=self.current_row, column=1)
        title_cell.value = f"Daily Report - {self.date_str}"
        title_cell.font = Font(name="Calibri", size=14, bold=True)
        title_cell.alignment = ExcelStyles.CENTER_CENTER
        self.ws.row_dimensions[self.current_row].height = 25
        self.current_row += 2

    def _add_collections_section(self):
        """Add collections table"""
        if not self.day_data["collections"]:
            return

        # Section header
        self.ws.merge_cells(f"A{self.current_row}:E{self.current_row}")
        section_cell = self.ws.cell(row=self.current_row, column=1, value="COLLECTIONS")
        ExcelStyles.apply_style(section_cell, ExcelStyles.pattern_4_section())
        self.current_row += 1

        # Table headers
        headers = [
            "Card No",
            "Procedure",
            "Payment Method",
            "Invoice Source",
            "Amount",
        ]
        self._write_header_row(headers)

        # Data rows
        for collection in self.day_data["collections"]:
            row_data = [
                collection["card_no"],
                collection["procedure"],
                collection["payment_method"],
                collection.get("invoice_source", ""),
                collection["amount"],
            ]
            self._write_data_row(row_data, amount_col=5)

        self.current_row += 1  # Spacer

    def _add_expenses_section(self):
        """Add expenses table"""
        if not self.day_data["expenses"]:
            return

        # Section header
        self.ws.merge_cells(f"A{self.current_row}:D{self.current_row}")
        section_cell = self.ws.cell(row=self.current_row, column=1, value="EXPENSES")
        ExcelStyles.apply_style(section_cell, ExcelStyles.pattern_4_section())
        self.current_row += 1

        # Table headers
        headers = [
            "Expense Name",
            "Payment Method",
            "Amount",
        ]
        self._write_header_row(headers)

        # Data rows
        for expense in self.day_data["expenses"]:
            row_data = [
                expense["expense_name"],
                expense["payment_method"],
                expense["amount"],
            ]
            self._write_data_row(row_data, amount_col=3)

        self.current_row += 1  # Spacer

    def _add_totals_section(self):
        """Add totals section"""
        self.current_row += 1

        # Section header
        self.ws.merge_cells(f"A{self.current_row}:D{self.current_row}")
        totals_header = self.ws.cell(
            row=self.current_row, column=1, value="DAILY TOTALS"
        )
        ExcelStyles.apply_style(totals_header, ExcelStyles.pattern_4_section())
        self.current_row += 1

        totals = self.day_data["totals"]

        # Payment method totals
        payment_totals = [
            ("Cash Total", totals["cash_total"]),
            ("Mpesa Total", totals["mpesa_total"]),
            ("Till Total", totals["till_total"]),
            ("Invoice Total", totals["invoice_total"]),
            ("Card Total", totals["card_total"]),
            ("Mobile Money Total", totals["mobile_money_total"]),
        ]

        for label, value in payment_totals:
            self._write_total_row(label, value, bold=False)

        self.current_row += 1  # Spacer

        # Summary totals
        summary_totals = [
            ("Gross Collections", totals["gross_collections"]),
            ("Total Expenses", totals["total_expenses"]),
        ]

        for label, value in summary_totals:
            self._write_total_row(label, value, bold=True)

        self.current_row += 1  # Spacer

        # Net total
        self._write_total_row("Net Total", totals["net_total"], bold=True)

    def _write_header_row(self, headers):
        """Write a header row with Pattern 1 styling"""
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.current_row, column=col, value=header)
            ExcelStyles.apply_style(cell, ExcelStyles.pattern_1_header())
        self.current_row += 1

    def _write_data_row(self, row_data, amount_col=None):
        """Write a data row with Pattern 3 styling"""
        for col, value in enumerate(row_data, 1):
            cell = self.ws.cell(row=self.current_row, column=col, value=value)
            if col == amount_col:
                ExcelStyles.apply_style(cell, ExcelStyles.pattern_3_data_right())
            else:
                ExcelStyles.apply_style(cell, ExcelStyles.pattern_3_data())
        self.current_row += 1

    def _write_total_row(self, label, value, bold=False):
        """Write a total row with Pattern 5 styling"""
        label_cell = self.ws.cell(row=self.current_row, column=3, value=label)
        value_cell = self.ws.cell(row=self.current_row, column=4, value=value)

        if bold:
            ExcelStyles.apply_style(label_cell, ExcelStyles.pattern_5_totals_bold())
            ExcelStyles.apply_style(value_cell, ExcelStyles.pattern_5_totals_bold())
        else:
            ExcelStyles.apply_style(label_cell, ExcelStyles.pattern_5_totals())
            ExcelStyles.apply_style(value_cell, ExcelStyles.pattern_5_totals())

        value_cell.alignment = ExcelStyles.RIGHT_CENTER
        self.current_row += 1

    def _set_column_widths(self):
        """Set appropriate column widths"""
        widths = {"A": 25, "B": 18, "C": 18, "D": 15}
        for col, width in widths.items():
            self.ws.column_dimensions[col].width = width

    def _save_to_bytes(self):
        """Save workbook to BytesIO"""
        bio = BytesIO()
        self.wb.save(bio)
        bio.seek(0)
        return bio
